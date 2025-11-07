"""
Parsers
"""
import csv
import json

from openpyxl import load_workbook
from openpyxl.cell import Cell

from .constants import *

try:
    from model import *
except:
    from pySeatShuffle.model import *


class PeopleParser:
    """
    ### People Table Format
    File Format: csv
    with table head: name,prop_name_1,prop_name_2,...
    """

    def __init__(self):
        pass

    def get_keys(self, file_path: str):
        with open(file_path, 'r') as f:
            reader = csv.reader(f)
            head = next(reader)
            for row in reader:
                properties = {}
                for i in range(0, len(row)):
                    properties[head[i]] = row[i]
                return list(properties.keys())
            return None

    # noinspection PyMethodMayBeStatic
    def parse(self, file_path: str, key: str):
        people = []
        with open(file_path, 'r') as f:
            reader = csv.reader(f)
            head = next(reader)
            id = 1
            for row in reader:
                properties = {}
                for i in range(0, len(row)):
                    properties[head[i]] = row[i]
                people.append(Person(id, properties, key))
                id += 1
        return people


class SeatTableParser:
    # noinspection PyMethodMayBeStatic
    def parse(self, file_path):
        return None


class SeatTableParserXlsx(SeatTableParser):
    def __init__(self):
        super().__init__()

    def parse(self, file_path):
        metadata = SeatTableMetadataXlsx(file_path)

        wb = load_workbook(file_path)
        ws = wb.worksheets[0]
        visited = set()  # 记录已访问的座位单元格坐标 (row, column)
        seat_groups = []
        gen_time_cells = []

        # 第一次遍历，收集所有占位符单元格，找到座位表offset
        offset_col = -1
        offset_row = -1
        max_row = 0
        max_col = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == XLSX_SEAT_PLACEHOLDER:
                    if offset_col == -1 or offset_col > cell.column:
                        offset_col = cell.column
                    if offset_row == -1 or offset_row > cell.row:
                        offset_row = cell.row
                    if cell.row > max_row:
                        max_row = cell.row
                    if cell.column > max_col:
                        max_col = cell.column
                if cell.value == XLSX_GEN_TIME_PLACEHOLDER:
                    gen_time_cells.append(cell)
        max_row -= offset_row - 1
        max_col -= offset_col - 1

        metadata.offset_row = offset_row - 1  # don't know this
        metadata.offset_col = offset_col - 1

        # 处理生成时间占位符
        if gen_time_cells:
            gen_time_cell = gen_time_cells[0]
            metadata.gen_time_cell_pos = (gen_time_cell.row, gen_time_cell.column)
        else:
            metadata.gen_time_cell_pos = None

        # 第二次遍历，处理座位占位符
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == XLSX_SEAT_PLACEHOLDER and (cell.row, cell.column) not in visited:
                    # 使用洪水填充找到连续区域
                    region = self._find_contiguous_region(cell, ws, visited)
                    # 检查区域四周是否有Thin边框
                    if self._is_region_bordered(region, ws):
                        seats = []
                        for region_cell in region:
                            # 将单元格位置转换为坐标，r为行，c为列
                            r = region_cell.row - offset_row + 1
                            c = region_cell.column - offset_col + 1
                            seats.append(Seat((r, c), name=None))  # 可根据需要添加名称
                        seat_group = SeatGroup(seats, name=None)  # 组名可后续处理
                        seat_groups.append(seat_group)

        # 假设表格尺寸为最大行列
        return SeatTable(seat_groups, size=(max_row, max_col), metadata=metadata)

    def _find_contiguous_region(self, start_cell: Cell, ws, visited: set):
        """使用BFS找到与start_cell相连的连续区域，区域内的单元格共享非Thin边框"""
        queue = [start_cell]
        region = []
        while queue:
            cell = queue.pop(0)
            coord = (cell.row, cell.column)
            if coord in visited:
                continue
            visited.add(coord)
            region.append(cell)
            # 检查四个方向
            directions = [('left', 0, -1), ('right', 0, 1),
                          ('top', -1, 0), ('bottom', 1, 0)]
            for border_dir, dr, dc in directions:
                adj_row = cell.row + dr
                adj_col = cell.column + dc
                if 1 <= adj_row <= ws.max_row and 1 <= adj_col <= ws.max_column:
                    adj_cell = ws.cell(row=adj_row, column=adj_col)
                    if adj_cell.value == XLSX_SEAT_PLACEHOLDER:
                        # 检查当前单元格与相邻单元格之间的边框是否为非Thin
                        if self._is_border_open(cell, border_dir):
                            if (adj_row, adj_col) not in visited:
                                queue.append(adj_cell)
        return region

    def _is_border_open(self, cell: Cell, direction: str) -> bool:
        """检查指定方向的边框是否为非Thin样式"""
        border = getattr(cell.border, direction)
        return border.style is None or border.style != 'thin'

    def _is_region_bordered(self, region: list[Cell], ws) -> bool:
        """检查区域四周是否全部由Thin边框包围"""
        if not region:
            return False
        # 获取区域的边界
        rows = {c.row for c in region}
        cols = {c.column for c in region}
        min_row, max_row = min(rows), max(rows)
        min_col, max_col = min(cols), max(cols)

        # 检查左边界（所有行的min_col列的左边框是否为thin）
        for row in rows:
            cell = ws.cell(row=row, column=min_col)
            if cell.border.left.style != 'thin':
                return False
        # 检查右边界（所有行的max_col列的右边框）
        for row in rows:
            cell = ws.cell(row=row, column=max_col)
            if cell.border.right.style != 'thin':
                return False
        # 检查上边界（所有列的min_row行的上边框）
        for col in cols:
            cell = ws.cell(row=min_row, column=col)
            if cell.border.top.style != 'thin':
                return False
        # 检查下边界（所有列的max_row行的下边框）
        for col in cols:
            cell = ws.cell(row=max_row, column=col)
            if cell.border.bottom.style != 'thin':
                return False
        return True


class SeatTableParserJson(SeatTableParser):
    """
    ### Seat Table Format
    File Format: json
    {
        "name": "smth",  (optional)
        "size": [w, h],
        "seat_groups": [
            {
                "name": "smth",  (optional)
                "seats": [
                    {
                        name: "smth",  (optional)
                        user: "smth",
                        pos: [x, y]
                    },
                    ...
                ]
            },
            ...
        ]
    }
    """

    # noinspection PyMethodMayBeStatic
    def parse(self, file_path):
        metadata = SeatTableMetadataJson(file_path)

        with open(file_path, 'r') as f:
            data = json.load(f)
        seat_groups = []
        for group in data['seat_groups']:
            seats = []
            for seat in group['seats']:
                seats.append(Seat(seat['pos'], seat.get('name', None)))
            seat_groups.append(SeatGroup(seats, group.get('name', None)))
        return SeatTable(seat_groups, data.get('size', None), data.get('name', None), metadata=metadata)


class RulesetParser:
    """
    ### Ruleset Format
    File Format: json
    {
        "rules": [
            {
                "t": "smth",
                "prop": [
                    "smth",
                    "smth",
                    ...
                ],
                "reversed": true  (optional)
            },
            ...
        ],
        "relations": WIP
    """

    def __init__(self):
        pass

    # noinspection PyMethodMayBeStatic
    def parse(self, file_path):
        with open(file_path, 'r') as f:
            data = json.load(f)
        rules = []
        for _rule in data['rules']:
            rules.append(Rule(_rule['t'], _rule['prop'], _rule.get('reversed', False)))
        return Ruleset(rules, data.get('relations', None))


default_people_parser = PeopleParser()
default_seat_table_parser_json = SeatTableParserJson()
default_seat_table_parser_xlsx = SeatTableParserXlsx()
default_ruleset_parser = RulesetParser()


def parse_people(file_path, key: str):
    return default_people_parser.parse(file_path, key)


def parse_seat_table(file_path):
    if file_path.endswith('.json'):
        return default_seat_table_parser_json.parse(file_path)
    elif file_path.endswith('.xlsx'):
        return default_seat_table_parser_xlsx.parse(file_path)
    else:
        raise ValueError('Unsupported file format')


def parse_ruleset(file_path):
    return default_ruleset_parser.parse(file_path)
