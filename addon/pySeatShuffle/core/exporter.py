import datetime
import json

from openpyxl.styles import NamedStyle

from model import *
from .constants import *

import zbToolLib as zbt


class SeatTableExporter:
    def export(self, table: SeatTable, format: str = F_JSON, path: str = None, template: str = None, create_dir=True):
        """
        Export a specific seat table in a specific format
        :param create_dir:
        :param table: a SeatTable object
        :param format: "F_" started constants, default follows the one specified in metadata, if None then json
        :param path: if None, then use the filename stored in metadata, and add "_export"
        :param template: exclusively for F_XLSX alike format
        """
        if format == F_JSON:
            self.export_json(table, path, create_dir)
        elif format == F_XLSX:
            self.export_xlsx(table, path, template, create_dir)
        elif format == F_XLS:
            self.export_xlsx(table, path, template, create_dir, use_xls=True)
        else:
            if hasattr(table.metadata, "format") and table.metadata.format in [F_JSON, F_XLSX, F_XLS]:
                self.export(table, table.metadata.format, path, template)
            else:
                self.export(table, F_JSON, path)


    def export_json(self, table: SeatTable, path, create_dir):
        path = self.generate_path(table, path, ".json")
        if create_dir and not zbt.existPath(zbt.getFileDir(path)):
            zbt.createDir(zbt.getFileDir(path))
        with open(path, "w", encoding="utf-8") as f:
            json.dump(table.to_dict(), f)

    def export_xlsx(self, table: SeatTable, path, template, create_dir, use_xls=False):
        path = self.generate_path(table, path, ".xlsx" if not use_xls else ".xls")

        if create_dir and not zbt.existPath(zbt.getFileDir(path)):
            zbt.createDir(zbt.getFileDir(path))

        metadata = table.metadata

        if template is None:
            if (metadata is not None and
                    hasattr(metadata, 'format') and
                    metadata.format == F_XLSX and
                    hasattr(metadata, 'file_path')
            ):
                template = metadata.file_path
            else:
                raise ValueError("Can't find template for exporting!")

        if template is None:
            raise ValueError("Template file path is not specified!")

        from openpyxl import load_workbook
        wb = load_workbook(template)
        ws = wb.active

        if hasattr(metadata, "gen_time_cell_pos") and metadata.gen_time_cell_pos is not None:
            time_cell = ws.cell(*metadata.gen_time_cell_pos)
            time_cell.value = datetime.datetime.now()
            time_cell.style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")

        if hasattr(metadata, "offset_row") and metadata.offset_row is not None:
            offset_row = metadata.offset_row
        else:
            offset_row = 0

        if hasattr(metadata, "offset_col") and metadata.offset_col is not None:
            offset_col = metadata.offset_col
        else:
            offset_col = 0

        for seat_group in table.get_seat_groups():
            for seat in seat_group.get_seats():
                cell = ws.cell(seat.get_pos()[0] + offset_row, seat.get_pos()[1] + offset_col)
                if cell.value == XLSX_SEAT_PLACEHOLDER:
                    cell.value = ""
                if seat.get_user() is not None:
                    ws.cell(seat.get_pos()[0] + offset_row, seat.get_pos()[1] + offset_col).value = seat.get_user().get_name()

        wb.save(path)

    # noinspection PyMethodMayBeStatic
    def generate_path(self, table, path, suffix):
        if path is None:
            metadata = table.metadata
            if metadata is not None and hasattr(metadata, 'file_path'):
                file_path = metadata.file_path
                path = zbt.getFileDir(file_path) + "/" + zbt.getFileName(file_path, False) + "_export"
                if zbt.existPath(path + suffix):
                    i = 1
                    while zbt.existPath(path + " (" + str(i) + ")" + suffix):
                        i += 1
                    path = path + " (" + str(i) + ")" + suffix
                else:
                    path = path + suffix
        return path


default_seat_table_exporter = SeatTableExporter()
